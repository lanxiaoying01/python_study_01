# author Amelia
# 2023/8/7 15:09
import os
from openpyxl import load_workbook
from django.shortcuts import render,redirect,HttpResponse
from app01 import models

def dept_list(request):
    """ 部门列表 """
    # 去数据库获取部门列表
    queryset = models.Department.objects.all()
    return render(request,'dept_list.html',{'queryset': queryset})

def dept_add(request):
    """ 添加部门 """
    if request.method=='GET':
        return render(request, 'dept_add.html')
    else:
        title = request.POST.get("title")
        models.Department.objects.create(title=title)
        return redirect('/dept/list')

def dept_delete(request):
    """ 删除 """
    nid = request.GET.get('nid')
    models.Department.objects.filter(id=nid).delete()
    return redirect('/dept/list')

def dept_edit(request,nid):
    """ 修改部门 """
    if request.method=='GET':
        row_obj = models.Department.objects.get(id=nid)
        return render(request, 'dept_edit.html',{'obj': row_obj})
    else:
        title = request.POST.get("title")
        models.Department.objects.filter(id=nid).update(title=title)
        return redirect('/dept/list')


def dept_multi(request):
    """ 批量删除 """
    # 1.获取用户上传的文件对象
    file_object = request.FILES.get("exc")

    # 2.对象传递给openpyxl,由openpyxl读取文件的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    # 3.循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        print(row)
        text = row[0].value
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)

    return redirect('/dept/list')

